from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from .models import Dht11, Incident, IncidentComment, ArchiveIncident, UserProfile, TemperatureThreshold
import csv
from .serializers import DHT11serialize
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.contrib.auth.models import User
import json
from django.core.mail import send_mail
from DHT.utils import send_telegram, send_whatsapp
from django.conf import settings
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from DHT.ntfy_notifications import send_ntfy_to_multiple_users

# ==================== HELPER FUNCTIONS ====================

def get_user_role(user):
    """Get user role safely"""
    if hasattr(user, 'profile'):
        return user.profile.role
    return 'visiteur'


# ==================== AUTHENTICATION ====================

def login_view(request):
    """Page de connexion"""
    if request.user.is_authenticated:
        return redirect('dashboard')

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            next_url = request.GET.get('next', 'dashboard')
            return redirect(next_url)
        else:
            messages.error(request, 'Nom d\'utilisateur ou mot de passe incorrect.')

    return render(request, 'auth/login.html')


def logout_view(request):
    """Déconnexion"""
    logout(request)
    return redirect('login')


@login_required
def profile_view(request):
    """Page de profil utilisateur"""
    return render(request, 'auth/profile.html')


# ==================== DASHBOARD ====================

@login_required
def dashboard(request):
    """Dashboard principal - page d'accueil"""
    return render(request, "dashboard.html")


# ==================== GRAPH PAGES ====================

@login_required
def graph_temp(request):
    """Page graphique température"""
    return render(request, "graph_temp.html")


@login_required
def graph_hum(request):
    """Page graphique humidité"""
    return render(request, "graph_hum.html")


def graphique(request):
    """Ancienne page graphique - garde pour compatibilité"""
    data = Dht11.objects.all()
    return render(request, 'chart.html', {'data': data})


def table(request):
    """Ancienne page table - garde pour compatibilité"""
    derniere_ligne = Dht11.objects.last()
    if not derniere_ligne:
        return render(request, 'value.html', {'valeurs': None})

    derniere_date = derniere_ligne.dt
    delta_temps = timezone.now() - derniere_date
    difference_minutes = delta_temps.seconds // 60

    temps_ecoule = ' il y a ' + str(difference_minutes) + ' min'
    if difference_minutes > 60:
        temps_ecoule = ('il y a ' + str(difference_minutes // 60) + 'h ' +
                        str(difference_minutes % 60) + 'min')

    valeurs = {
        'date': temps_ecoule,
        'id': derniere_ligne.id,
        'temp': derniere_ligne.temp,
        'hum': derniere_ligne.hum
    }
    return render(request, 'value.html', {'valeurs': valeurs})


# ==================== DATA VIEWS ====================

def latest_json(request):
    """API - Dernière mesure en JSON"""
    last = Dht11.objects.order_by('-dt').values('temp', 'hum', 'dt').first()
    if not last:
        return JsonResponse({"detail": "Aucune donnée disponible"}, status=404)
    return JsonResponse({
        "temperature": last["temp"],
        "humidity": last["hum"],
        "timestamp": last["dt"].isoformat()
    })


def chart_data(request):
    """API - Toutes les données pour graphiques"""
    dht = Dht11.objects.all().order_by('dt')
    data = {
        'temps': [dt.dt.isoformat() for dt in dht],
        'temperature': [temp.temp for temp in dht],
        'humidity': [hum.hum for hum in dht]
    }
    return JsonResponse(data)


def chart_data_jour(request):
    """API - Données des dernières 24h"""
    now = timezone.now()
    last_24_hours = now - timezone.timedelta(hours=24)
    dht = Dht11.objects.filter(dt__range=(last_24_hours, now)).order_by('dt')
    data = {
        'temps': [dt.dt.isoformat() for dt in dht],
        'temperature': [temp.temp for temp in dht],
        'humidity': [hum.hum for hum in dht]
    }
    return JsonResponse(data)


def chart_data_semaine(request):
    """API - Données de la dernière semaine"""
    date_debut_semaine = timezone.now() - datetime.timedelta(days=7)
    dht = Dht11.objects.filter(dt__gte=date_debut_semaine).order_by('dt')
    data = {
        'temps': [dt.dt.isoformat() for dt in dht],
        'temperature': [temp.temp for temp in dht],
        'humidity': [hum.hum for hum in dht]
    }
    return JsonResponse(data)


def chart_data_mois(request):
    """API - Données du dernier mois"""
    date_debut_mois = timezone.now() - datetime.timedelta(days=30)
    dht = Dht11.objects.filter(dt__gte=date_debut_mois).order_by('dt')
    data = {
        'temps': [dt.dt.isoformat() for dt in dht],
        'temperature': [temp.temp for temp in dht],
        'humidity': [hum.hum for hum in dht]
    }
    return JsonResponse(data)


# ==================== MANUAL DATA ENTRY - FIXED ====================

@csrf_exempt
@require_http_methods(["POST"])
def manual_data_entry(request):
    """Entrée manuelle de données - FIXED LOGIC"""
    try:
        data = json.loads(request.body)
        temperature = float(data.get('temp', 0))
        humidity = float(data.get('hum', 0))

        print(f"\n{'=' * 60}")
        print(f"📊 MANUAL ENTRY: Temp={temperature}°C, Hum={humidity}%")

        # Créer l'entrée DHT11
        dht_entry = Dht11.objects.create(temp=temperature, hum=humidity)

        # Get temperature thresholds
        threshold = TemperatureThreshold.objects.first()
        if not threshold:
            threshold = TemperatureThreshold.objects.create(min_temp=2.0, max_temp=8.0)

        min_temp = threshold.min_temp
        max_temp = threshold.max_temp

        print(f"🎯 THRESHOLDS: Min={min_temp}°C, Max={max_temp}°C")

        # Chercher incident actif
        incident = Incident.objects.filter(actif=True).first()
        now = timezone.now()

        # Check if temperature is OUTSIDE the normal range (ABNORMAL)
        is_abnormal = (temperature < min_temp) or (temperature > max_temp)

        print(f"🔍 Check: {temperature}°C is {'ABNORMAL' if is_abnormal else 'NORMAL'}")

        if is_abnormal:
            # TEMPERATURE ABNORMAL - CREATE OR INCREMENT
            print(f"🚨 ABNORMAL - Creating/incrementing incident")

            if incident:
                # Incrémenter compteur jusqu'à 9
                if incident.compteur < 9:
                    if (not incident.last_increment) or ((now - incident.last_increment).total_seconds() >= 10):
                        incident.compteur += 1
                        incident.last_increment = now
                        incident.temperature = temperature
                        incident.humidity = humidity
                        incident.save()
                        print(f"✅ Incremented to {incident.compteur}/9")
            else:
                # Créer nouvel incident
                incident = Incident.objects.create(
                    actif=True,
                    compteur=1,
                    date_debut=now,
                    last_increment=now,
                    temperature=temperature,
                    humidity=humidity,
                    status='en_cours'
                )
                print(f"✅ New incident created ID={incident.id}")

            message = f"⚠️ Alerte!\nTemp: {temperature:.1f}°C (range: {min_temp}-{max_temp}°C)\nHum: {humidity:.1f}%\nCompteur: {incident.compteur}/9"
            try:
                send_ntfy_to_multiple_users(
                    temperature=temperature,
                    humidity=humidity,
                    incident_count=incident.compteur
                )
                print("✅ Ntfy notification sent for manual entry")
            except Exception as e:
                print(f"❌ Error sending ntfy: {e}")

            try:
                send_mail(
                    subject="Alerte Température",
                    message=message,
                    from_email=settings.EMAIL_HOST_USER,
                    recipient_list=["imanejennane23@gmail.com"],
                    fail_silently=False,
                )
            except Exception as e:
                print(f"Email error: {e}")

            try:
                send_telegram(message)
            except Exception as e:
                print(f"Telegram error: {e}")

            try:
                send_whatsapp(message)
            except Exception as e:
                print(f"WhatsApp error: {e}")

        else:
            # TEMPERATURE NORMAL - CLOSE INCIDENT
            print(f"✅ NORMAL - Closing incident if exists")

            if incident:
                # Archiver
                ArchiveIncident.objects.create(
                    date_debut=incident.date_debut,
                    date_fin=now,
                    compteur=incident.compteur,
                    status='termine',
                    temperature=incident.temperature,
                    humidity=incident.humidity,
                    nom_op1=incident.nom_op1,
                    op1_checked=incident.op1_checked,
                    op1_comment=incident.op1_comment or '',
                    op1_operateur_name=incident.op1_operateur.username if incident.op1_operateur else '',
                    op1_date=incident.op1_date,
                    nom_op2=incident.nom_op2,
                    op2_checked=incident.op2_checked,
                    op2_comment=incident.op2_comment or '',
                    op2_operateur_name=incident.op2_operateur.username if incident.op2_operateur else '',
                    op2_date=incident.op2_date,
                    nom_op3=incident.nom_op3,
                    op3_checked=incident.op3_checked,
                    op3_comment=incident.op3_comment or '',
                    op3_operateur_name=incident.op3_operateur.username if incident.op3_operateur else '',
                    op3_date=incident.op3_date,
                )

                # Fermer
                incident.actif = False
                incident.date_fin = now
                incident.status = 'termine'
                incident.last_increment = None
                incident.save()

                print(f"✅ Incident #{incident.id} closed and archived")

        print(f"{'=' * 60}\n")

        return JsonResponse({
            'success': True,
            'message': 'Données enregistrées avec succès',
            'id': dht_entry.id,
            'temperature': temperature,
            'humidity': humidity
        })

    except ValueError:
        return JsonResponse({'success': False, 'error': 'Valeurs invalides'}, status=400)
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ==================== INCIDENT STATUS ====================

@login_required
def incident_status(request):
    """API - Statut incident actuel"""
    incident = Incident.objects.filter(actif=True).first()

    user_role = get_user_role(request.user)

    can_edit_op1 = user_role in ['admin', 'operateur1']
    can_edit_op2 = user_role in ['admin', 'operateur2']
    can_edit_op3 = user_role in ['admin', 'operateur3']
    can_comment = user_role != 'visiteur'
    can_accuse_reception = user_role != 'visiteur'

    if incident:
        comments = incident.comments.all().order_by('created_at')
        comments_data = [{
            'author': comment.author.username,
            'content': comment.content,
            'created_at': comment.created_at.isoformat()
        } for comment in comments]

        return JsonResponse({
            "incident_actif": True,
            "id": incident.id,
            "compteur": incident.compteur,
            "date_debut": incident.date_debut.isoformat(),
            "temperature": incident.temperature,
            "humidity": incident.humidity,
            "accuse_reception": incident.accuse_reception,
            "accuse_reception_operateur": incident.accuse_reception_operateur.username if incident.accuse_reception_operateur else None,
            "accuse_reception_date": incident.accuse_reception_date.isoformat() if incident.accuse_reception_date else None,
            "nom_op1": incident.nom_op1,
            "op1_checked": incident.op1_checked,
            "op1_comment": incident.op1_comment or "",
            "op1_operateur": incident.op1_operateur.username if incident.op1_operateur else None,
            "op1_date": incident.op1_date.isoformat() if incident.op1_date else None,
            "nom_op2": incident.nom_op2,
            "op2_checked": incident.op2_checked,
            "op2_comment": incident.op2_comment or "",
            "op2_operateur": incident.op2_operateur.username if incident.op2_operateur else None,
            "op2_date": incident.op2_date.isoformat() if incident.op2_date else None,
            "nom_op3": incident.nom_op3,
            "op3_checked": incident.op3_checked,
            "op3_comment": incident.op3_comment or "",
            "op3_operateur": incident.op3_operateur.username if incident.op3_operateur else None,
            "op3_date": incident.op3_date.isoformat() if incident.op3_date else None,
            "comments": comments_data,
            "permissions": {
                "user_role": user_role,
                "can_edit_op1": can_edit_op1,
                "can_edit_op2": can_edit_op2,
                "can_edit_op3": can_edit_op3,
                "can_comment": can_comment,
                "can_accuse_reception": can_accuse_reception
            }
        })
    else:
        return JsonResponse({
            "incident_actif": False,
            "compteur": 0,
            "permissions": {
                "user_role": user_role,
                "can_edit_op1": can_edit_op1,
                "can_edit_op2": can_edit_op2,
                "can_edit_op3": can_edit_op3,
                "can_comment": can_comment,
                "can_accuse_reception": can_accuse_reception
            }
        })


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def update_incident(request):
    """API - Mettre à jour incident"""
    incident = Incident.objects.filter(actif=True).first()
    if not incident:
        return JsonResponse({"error": "Aucun incident actif"}, status=400)

    user_role = get_user_role(request.user)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    if 'accuse_reception' in data:
        if user_role == 'visiteur':
            return JsonResponse({"error": "Permission refusée"}, status=403)
        incident.accuse_reception = data['accuse_reception']
        if data['accuse_reception']:
            incident.accuse_reception_operateur = request.user
            incident.accuse_reception_date = timezone.now()

    if 'op1_checked' in data or 'op1_comment' in data:
        if user_role not in ['admin', 'operateur1']:
            return JsonResponse({"error": "Permission refusée pour op1"}, status=403)
        if 'op1_checked' in data:
            incident.op1_checked = data['op1_checked']
            if data['op1_checked']:
                incident.op1_operateur = request.user
                incident.op1_date = timezone.now()
        if 'op1_comment' in data:
            incident.op1_comment = data['op1_comment']

    if 'op2_checked' in data or 'op2_comment' in data:
        if user_role not in ['admin', 'operateur2']:
            return JsonResponse({"error": "Permission refusée pour op2"}, status=403)
        if 'op2_checked' in data:
            incident.op2_checked = data['op2_checked']
            if data['op2_checked']:
                incident.op2_operateur = request.user
                incident.op2_date = timezone.now()
        if 'op2_comment' in data:
            incident.op2_comment = data['op2_comment']

    if 'op3_checked' in data or 'op3_comment' in data:
        if user_role not in ['admin', 'operateur3']:
            return JsonResponse({"error": "Permission refusée pour op3"}, status=403)
        if 'op3_checked' in data:
            incident.op3_checked = data['op3_checked']
            if data['op3_checked']:
                incident.op3_operateur = request.user
                incident.op3_date = timezone.now()
        if 'op3_comment' in data:
            incident.op3_comment = data['op3_comment']

    incident.save()
    return JsonResponse({"success": True})


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def add_incident_comment(request, incident_id):
    """Ajouter un commentaire"""
    user_role = get_user_role(request.user)

    if user_role == 'visiteur':
        return JsonResponse({'success': False, 'error': 'Permission refusée'}, status=403)

    try:
        incident = Incident.objects.get(id=incident_id, actif=True)
        data = json.loads(request.body)
        content = data.get('content', '').strip()

        if not content:
            return JsonResponse({'success': False, 'error': 'Commentaire vide'}, status=400)

        comment = IncidentComment.objects.create(
            incident=incident,
            author=request.user,
            content=content
        )

        return JsonResponse({
            'success': True,
            'comment': {
                'id': comment.id,
                'author': request.user.username,
                'content': comment.content,
                'created_at': comment.created_at.isoformat()
            }
        })

    except Incident.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Incident non trouvé'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def check_create_incident(request):
    """Vérifier et créer/mettre à jour un incident"""
    try:
        data = json.loads(request.body)
        temperature = float(data.get('temperature', 0))

        threshold = TemperatureThreshold.objects.first()
        if not threshold:
            threshold = TemperatureThreshold.objects.create(min_temp=2.0, max_temp=8.0)

        min_temp = threshold.min_temp
        max_temp = threshold.max_temp

        if temperature < min_temp or temperature > max_temp:
            incident = Incident.objects.filter(actif=True).first()
            now = timezone.now()

            if incident:
                if incident.compteur < 9:
                    if (not incident.last_increment) or ((now - incident.last_increment).total_seconds() >= 10):
                        incident.compteur += 1
                        incident.last_increment = now
                        incident.save()

                return JsonResponse({
                    'success': True,
                    'message': 'Incident mis à jour',
                    'compteur': incident.compteur,
                    'created': False
                })
            else:
                incident = Incident.objects.create(
                    actif=True,
                    compteur=1,
                    date_debut=now,
                    last_increment=now,
                    temperature=temperature
                )

                return JsonResponse({
                    'success': True,
                    'message': 'Incident créé',
                    'compteur': incident.compteur,
                    'created': True
                })
        else:
            incident = Incident.objects.filter(actif=True).first()
            if incident:
                incident.actif = False
                incident.date_fin = timezone.now()
                incident.last_increment = None
                incident.save()

                return JsonResponse({
                    'success': True,
                    'message': 'Incident fermé',
                    'closed': True
                })

            return JsonResponse({
                'success': True,
                'message': 'Pas d\'incident à créer'
            })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# ==================== ARCHIVE - FIXED ====================

@login_required
def archive_incidents(request):
    """Page archive - FIXED to show ALL past incidents"""
    try:
        # Get ALL closed incidents from both tables
        closed_incidents = Incident.objects.filter(actif=False).order_by('-date_debut')
        archived_incidents = ArchiveIncident.objects.all().order_by('-date_debut')

        # Combine into single list
        all_archives = list(closed_incidents) + list(archived_incidents)

        # Sort by date
        all_archives.sort(key=lambda x: x.date_debut, reverse=True)

        print(
            f"📦 Archive page: Found {len(all_archives)} total incidents ({len(closed_incidents)} closed, {len(archived_incidents)} archived)")

        return render(request, 'archives_incidents.html', {'archives': all_archives})
    except Exception as e:
        print(f"❌ Error in archive_incidents: {e}")
        import traceback
        traceback.print_exc()
        return render(request, 'archives_incidents.html', {'archives': [], 'error': str(e)})


@login_required
def archive_incident_detail(request, incident_id):
    """Page détail - FIXED"""
    try:
        # Try Archive first
        incident = ArchiveIncident.objects.filter(id=incident_id).first()

        # If not found, try closed Incidents
        if not incident:
            incident = Incident.objects.filter(id=incident_id).first()

        if not incident:
            messages.error(request, 'Incident non trouvé')
            return redirect('archive_incidents')

        return render(request, 'archive_incident_detail.html', {'incident': incident})
    except Exception as e:
        print(f"❌ Error: {e}")
        messages.error(request, f'Erreur: {str(e)}')
        return redirect('archive_incidents')


# ==================== DOWNLOAD/EXPORT ====================

def download_csv(request):
    """Téléchargement CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="dht11_data.csv"'

    writer = csv.writer(response)
    writer.writerow(['ID', 'Temperature (°C)', 'Humidite (%)', 'Date et Heure'])

    model_values = Dht11.objects.values_list('id', 'temp', 'hum', 'dt')
    for row in model_values:
        writer.writerow(row)

    return response


def download_incidents_excel(request):
    """Téléchargement Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Incidents DHT11"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = [
        'ID', 'Date Début', 'Date Fin', 'Compteur', 'Statut',
        'Opération 1', 'Commentaire Op1',
        'Opération 2', 'Commentaire Op2',
        'Opération 3', 'Commentaire Op3'
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    incidents = list(Incident.objects.all().order_by('-date_debut'))
    archives = list(ArchiveIncident.objects.all().order_by('-date_debut'))
    all_incidents = incidents + archives

    for row_num, incident in enumerate(all_incidents, 2):
        if hasattr(incident, 'actif'):
            ws.cell(row=row_num, column=1).value = incident.id
            statut = 'Actif' if incident.actif else 'Fermé'
        else:
            ws.cell(row=row_num, column=1).value = f"A{incident.id}"
            statut = 'Archivé'

        ws.cell(row=row_num, column=2).value = incident.date_debut.strftime('%d/%m/%Y %H:%M:%S')
        date_fin = incident.date_fin.strftime('%d/%m/%Y %H:%M:%S') if incident.date_fin else ''
        ws.cell(row=row_num, column=3).value = date_fin
        ws.cell(row=row_num, column=4).value = incident.compteur
        ws.cell(row=row_num, column=5).value = statut

        ws.cell(row=row_num, column=6).value = '✓' if incident.op1_checked else '✗'
        ws.cell(row=row_num, column=7).value = incident.op1_comment or ''
        ws.cell(row=row_num, column=8).value = '✓' if incident.op2_checked else '✗'
        ws.cell(row=row_num, column=9).value = incident.op2_comment or ''
        ws.cell(row=row_num, column=10).value = '✓' if incident.op3_checked else '✗'
        ws.cell(row=row_num, column=11).value = incident.op3_comment or ''

        for col in range(1, 12):
            ws.cell(row=row_num, column=col).border = border

    column_widths = {'A': 8, 'B': 20, 'C': 20, 'D': 12, 'E': 12, 'F': 15, 'G': 30, 'H': 15, 'I': 30, 'J': 15, 'K': 30}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = 'A2'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response[
        'Content-Disposition'] = f'attachment; filename="incidents_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'

    wb.save(response)
    return response

# ==================== ADMIN PANEL ====================

@login_required
def admin_panel(request):
    """Panneau d'administration (Admin seulement)"""
    user_role = get_user_role(request.user)

    if user_role != 'admin':
        messages.error(request, 'Accès non autorisé.')
        return redirect('dashboard')

    action = request.GET.get('action', 'dashboard')

    if action == 'create_operateur' and request.method == 'POST':
        try:
            username = request.POST.get('username')
            password = request.POST.get('password')
            email = request.POST.get('email', '')
            full_name = request.POST.get('full_name', '')
            phone_number = request.POST.get('phone_number', '')
            role = request.POST.get('role')

            if User.objects.filter(username=username).exists():
                messages.error(request, 'Nom d\'utilisateur déjà utilisé')
            else:
                user = User.objects.create_user(username=username, password=password, email=email)
                UserProfile.objects.create(
                    user=user,
                    role=role,
                    full_name=full_name,
                    phone_number=phone_number,
                    email=email
                )
                messages.success(request, f'Opérateur {username} créé avec succès')

        except Exception as e:
            messages.error(request, f'Erreur: {str(e)}')

        return redirect('admin_panel')
    elif action == 'update_thresholds' and request.method == 'POST':
        try:
            min_temp = float(request.POST.get('min_temp'))
            max_temp = float(request.POST.get('max_temp'))

            if min_temp >= max_temp:
                messages.error(request, 'La température minimum doit être inférieure à la maximum')
            else:
                threshold = TemperatureThreshold.objects.first()
                if threshold:
                    threshold.min_temp = min_temp
                    threshold.max_temp = max_temp
                    threshold.updated_by = request.user
                    threshold.save()
                    messages.success(request, 'Seuils mis à jour avec succès')
                else:
                    TemperatureThreshold.objects.create(
                        min_temp=min_temp,
                        max_temp=max_temp,
                        updated_by=request.user
                    )
                    messages.success(request, 'Seuils créés avec succès')

        except Exception as e:
            messages.error(request, f'Erreur: {str(e)}')

        return redirect('admin_panel')

        # Get data for template
    operateurs = UserProfile.objects.all().order_by('user__username')
    threshold = TemperatureThreshold.objects.first()
    if not threshold:
        threshold = TemperatureThreshold.objects.create(min_temp=2.0, max_temp=8.0, updated_by=request.user)

    ROLE_CHOICES = UserProfile.ROLE_CHOICES

    context = {
        'operateurs': operateurs,
        'threshold': threshold,
        'ROLE_CHOICES': ROLE_CHOICES,
    }

    return render(request, 'admin_panel.html', context)